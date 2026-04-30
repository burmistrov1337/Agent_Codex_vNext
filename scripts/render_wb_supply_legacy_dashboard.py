from __future__ import annotations

import html
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path("D:/Agent_Codex_vNext")
SOURCE_XLSX = PROJECT_ROOT / "generated" / "marketplace" / "month_supply_2026-04-30" / "month_supply.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "generated" / "marketplace" / f"supply_plan_{date.today().isoformat()}"


@dataclass(slots=True)
class SupplyRow:
    sku: str
    nm_id: int
    vendor_code: str
    barcode: str
    price: float | None
    sales_30d: int
    sale_days_30d: int
    avg_daily: float
    seller_stock: int
    wb_stock: int
    coverage_days: float | None
    demand_30d: int
    safe_available: int
    ship_qty: int
    priority: str
    problem: str
    action: str


def main() -> None:
    rows = _read_rows(SOURCE_XLSX)
    rows.sort(key=lambda row: (row.ship_qty <= 0, -row.ship_qty, -row.sales_30d, row.sku.lower()))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    html_path = OUTPUT_DIR / "supply_dashboard_dark.html"
    xlsx_path = OUTPUT_DIR / "supply_plan.xlsx"
    md_path = OUTPUT_DIR / "supply_plan.md"
    template_path = OUTPUT_DIR / "template.xlsx"

    html_path.write_text(_render_dashboard(rows), encoding="utf-8")
    _write_plan_xlsx(xlsx_path, rows)
    _write_template_xlsx(template_path, [row for row in rows if row.ship_qty > 0])
    md_path.write_text(_render_markdown(rows), encoding="utf-8")

    print(f"output_dir={OUTPUT_DIR}")
    print(f"dashboard={html_path}")
    print(f"xlsx={xlsx_path}")
    print(f"md={md_path}")
    print(f"template={template_path}")
    print(f"rows={len(rows)}")
    print(f"ship_rows={sum(1 for row in rows if row.ship_qty > 0)}")
    print(f"ship_total={sum(row.ship_qty for row in rows if row.ship_qty > 0)}")


def _read_rows(path: Path) -> list[SupplyRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1))]
    index = {name: position for position, name in enumerate(headers)}
    rows: list[SupplyRow] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        rows.append(
            SupplyRow(
                sku=str(values[index["SKU"]] or ""),
                nm_id=int(values[index["Артикул WB"]] or 0),
                vendor_code=str(values[index["Артикул продавца"]] or ""),
                barcode=str(values[index["Баркод"]] or ""),
                price=_float_or_none(values[index["Цена WB после скидки продавца"]]),
                sales_30d=int(values[index["Продажи 30д"]] or 0),
                sale_days_30d=int(values[index["Дни продаж 30д"]] or 0),
                avg_daily=float(values[index["Средние продажи в день"]] or 0),
                seller_stock=int(values[index["Остаток продавца"]] or 0),
                wb_stock=int(values[index["WB-остаток"]] or 0),
                coverage_days=_float_or_none(values[index["Покрытие WB, дни"]]),
                demand_30d=int(values[index["Потребность на 30 дней"]] or 0),
                safe_available=int(values[index["Безопасно доступно к отгрузке"]] or 0),
                ship_qty=int(values[index["К отгрузке"]] or 0),
                priority=str(values[index["Приоритет"]] or ""),
                problem=str(values[index["Что не так"]] or ""),
                action=str(values[index["Что делать"]] or ""),
            )
        )
    return rows


def _float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _render_dashboard(rows: list[SupplyRow]) -> str:
    ship_rows = [row for row in rows if row.ship_qty > 0]
    watch_rows = [row for row in rows if row.ship_qty <= 0]
    generated_at = date.today().isoformat()
    summary_cards = [
        ("SKU с продажами и остатком", len(rows)),
        ("SKU к отгрузке", len(ship_rows)),
        ("Всего к отгрузке", sum(row.ship_qty for row in ship_rows)),
        ("Безопасно доступно", sum(row.safe_available for row in rows)),
    ]
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Поставка WB</title>
  <style>
    :root {{
      --bg: #07111f;
      --panel: #101a2a;
      --panel-2: #172437;
      --text: #f8fafc;
      --muted: #9fb2cc;
      --line: #2b3a50;
      --blue: #2978f0;
      --teal: #0f8f86;
      --amber: #31290d;
      --danger: #7f1d1d;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: var(--bg); color: var(--text); font-family: Segoe UI, Arial, sans-serif; }}
    .page {{ max-width: 1480px; margin: 0 auto; padding: 16px 14px 36px; }}
    .hero {{ padding: 30px 32px; border-radius: 22px; background: linear-gradient(135deg, #255fdf, #0d8078); box-shadow: 0 20px 50px rgba(0,0,0,.22); }}
    h1 {{ margin: 0 0 12px; font-size: 32px; }}
    h2 {{ margin: 28px 0 14px; font-size: 24px; }}
    p, li {{ color: var(--text); line-height: 1.45; }}
    .muted {{ color: var(--muted); }}
    .cards {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px; margin: 22px 0; }}
    .metric, .card {{ background: var(--panel); border-radius: 18px; padding: 20px; box-shadow: inset 0 0 0 1px rgba(255,255,255,.025); }}
    .metric .label {{ color: var(--muted); font-size: 13px; }}
    .metric .value {{ margin-top: 12px; font-size: 30px; font-weight: 800; }}
    .section {{ margin-top: 22px; }}
    table {{ width: 100%; border-collapse: separate; border-spacing: 0; overflow: hidden; border-radius: 16px; background: #0d1726; }}
    th, td {{ padding: 14px 16px; border-bottom: 1px solid var(--line); vertical-align: top; }}
    th {{ background: #1d2a3a; color: var(--text); text-align: left; font-weight: 800; cursor: pointer; user-select: none; }}
    th.sort-asc::after {{ content: " ▲"; color: #77e7dc; }}
    th.sort-desc::after {{ content: " ▼"; color: #77e7dc; }}
    td.num, th.num, .warehouse-table th:not(:first-child), .warehouse-table td:not(:first-child) {{ text-align: center; }}
    .print-controls {{ display: flex; justify-content: space-between; align-items: center; gap: 12px; flex-wrap: wrap; }}
    .warehouse-controls {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
    button, select {{ border: 0; border-radius: 13px; padding: 11px 14px; color: white; background: linear-gradient(135deg, var(--blue), var(--teal)); font-weight: 700; }}
    select {{ min-width: 360px; background: #142236; border: 1px solid var(--line); }}
    .remove-row-button {{ background: var(--danger); padding: 8px 10px; }}
    .editable {{ background: var(--amber); min-width: 110px; }}
    .priority-table.collapsed .priority-extra-row {{ display: none; }}
    .print-hide {{ display: table-cell; }}
    @media (max-width: 900px) {{
      .cards {{ grid-template-columns: 1fr; }}
      .page {{ padding: 10px; }}
      table {{ font-size: 13px; }}
      th, td {{ padding: 10px; }}
    }}
    @media print {{
      body.warehouse-print-mode {{ background: white !important; color: #111827 !important; }}
      body.warehouse-print-mode .page > * {{ display: none !important; }}
      body.warehouse-print-mode #warehouse-section {{ display: block !important; margin: 0 !important; }}
      body.warehouse-print-mode #warehouse-section .print-controls {{ display: none !important; }}
      body.warehouse-print-mode #warehouse-section table {{ background: white !important; color: #111827 !important; box-shadow: none !important; border-radius: 0 !important; }}
      body.warehouse-print-mode #warehouse-section th {{ background: #e5e7eb !important; color: #111827 !important; }}
      body.warehouse-print-mode #warehouse-section td, body.warehouse-print-mode #warehouse-section th {{ border-bottom: 1px solid #cbd5e1 !important; }}
      body.warehouse-print-mode #warehouse-section .editable {{ background: white !important; }}
      body.warehouse-print-mode #warehouse-section .print-hide {{ display: none !important; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <section class="hero">
      <h1>Поставка WB</h1>
      <p>Дата отчёта: {generated_at}. Период расчёта: последние 30 дней. Остаток продавца взят из seller inventory API. Города не распределяются, ТНВЭД в этот дашборд не включён.</p>
    </section>

    <section class="cards">
      {''.join(f"<div class='metric'><div class='label'>{html.escape(label)}</div><div class='value'>{value}</div></div>" for label, value in summary_cards)}
    </section>

    <section class="card">
      <h2>Сводка</h2>
      <ul>
        <li>В приоритете позиции, где текущий WB-остаток не закрывает месяц спроса.</li>
        <li>В складской таблице снизу уже лежат позиции, которые предлагается везти сейчас.</li>
        <li>Количество в складской таблице можно менять вручную; экспорт берёт именно ручной столбец «Количество».</li>
      </ul>
    </section>

    <section class="section">
      <div class="print-controls">
        <h2>Приоритет к отгрузке</h2>
        <button type="button" id="priority-toggle-button" onclick="togglePriorityRows()">Показать весь список</button>
      </div>
      {_render_priority_table(ship_rows)}
    </section>

    <section class="section">
      <h2>Не везти сейчас</h2>
      {_render_watch_table(watch_rows[:40])}
    </section>

    <section class="section card">
      <h2>Вывод и действия</h2>
      <ul>
        <li>Сначала отгружать SKU из таблицы приоритета с максимальным «К отгрузке» и нулевым или низким WB-остатком.</li>
        <li>Позиции из блока «Не везти сейчас» не включать в поставку без ручной причины: обычно потребность закрыта WB-остатком или расчётная безопасная отгрузка равна нулю.</li>
        <li>Перед передачей на склад проверь вручную столбец «Количество» и заполни сроки годности.</li>
      </ul>
    </section>

    <section class="section" id="warehouse-section">
      <h2>Таблица для склада</h2>
      <div class="print-controls">
        <div class="warehouse-controls">
          <select id="warehouse-product-select"></select>
          <button type="button" onclick="addSelectedProduct()">Добавить товар из анализа</button>
          <button type="button" onclick="exportWarehouseTemplate()">Выгрузить складскую таблицу</button>
        </div>
        <button type="button" onclick="printWarehouseTable()">Печать складской таблицы</button>
      </div>
      <table class="warehouse-table">
        <thead>
          <tr>
            <th>SKU</th>
            <th>Артикул WB</th>
            <th>Артикул продавца</th>
            <th>Баркод</th>
            <th class="print-hide">Остаток</th>
            <th class="print-hide">Остаток WB</th>
            <th class="print-hide">Цена со скидкой</th>
            <th class="print-hide">К отгрузке</th>
            <th>Количество</th>
            <th>Срок годности</th>
            <th class="print-hide">Убрать</th>
          </tr>
        </thead>
        <tbody id="warehouse-table-body">
          {''.join(_render_warehouse_row(row) for row in ship_rows)}
        </tbody>
      </table>
    </section>
  </div>
  <script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
  <script>
    const warehouseCatalog = [{','.join(_render_catalog_item(row) for row in rows)}];

    function getWarehouseBody() {{ return document.getElementById('warehouse-table-body'); }}
    function getSelectedNmIds() {{
      return Array.from(getWarehouseBody().querySelectorAll('tr[data-nm-id]')).map((row) => row.dataset.nmId);
    }}
    function refreshProductOptions() {{
      const select = document.getElementById('warehouse-product-select');
      const selected = new Set(getSelectedNmIds());
      const options = warehouseCatalog
        .filter((item) => !selected.has(item.nmId))
        .map((item) => `<option value="${{item.nmId}}">${{item.sku}}</option>`);
      select.innerHTML = options.length ? options.join('') : '<option value="">Все товары уже добавлены</option>';
      select.disabled = options.length === 0;
    }}
    function addSelectedProduct() {{
      const select = document.getElementById('warehouse-product-select');
      if (!select.value) return;
      const item = warehouseCatalog.find((entry) => entry.nmId === select.value);
      if (!item) return;
      const row = document.createElement('tr');
      row.dataset.nmId = item.nmId;
      row.innerHTML = `
        <td>${{item.sku}}</td>
        <td>${{item.nmId}}</td>
        <td>${{item.vendorCode}}</td>
        <td>${{item.barcode}}</td>
        <td class="print-hide">${{item.stock}}</td>
        <td class="print-hide">${{item.wbStock}}</td>
        <td class="print-hide">${{item.price}}</td>
        <td class="print-hide">${{item.shipQty}}</td>
        <td class="editable" contenteditable="true">${{item.shipQty}}</td>
        <td class="editable" contenteditable="true"></td>
        <td class="print-hide"><button class="remove-row-button" type="button" onclick="removeWarehouseItem(this)">Убрать</button></td>
      `;
      getWarehouseBody().appendChild(row);
      refreshProductOptions();
    }}
    function removeWarehouseItem(button) {{
      button.closest('tr').remove();
      refreshProductOptions();
    }}
    function printWarehouseTable() {{
      document.body.classList.add('warehouse-print-mode');
      window.print();
      window.setTimeout(() => document.body.classList.remove('warehouse-print-mode'), 300);
    }}
    function exportWarehouseTemplate() {{
      if (typeof XLSX === 'undefined') {{
        alert('Библиотека XLSX не загрузилась. Проверь интернет и повтори экспорт.');
        return;
      }}
      const rows = Array.from(getWarehouseBody().querySelectorAll('tr[data-nm-id]')).map((row) => {{
        const cells = row.querySelectorAll('td');
        return {{
          title: (cells[0]?.innerText || '').trim(),
          barcode: (cells[3]?.innerText || '').trim(),
          quantity: (cells[8]?.innerText || '').trim(),
        }};
      }}).filter((item) => item.title || item.barcode || item.quantity);
      const missing = rows.filter((item) => !item.quantity);
      if (missing.length) {{
        const preview = missing.slice(0, 5).map((item) => item.title || item.barcode || 'Без названия');
        const tail = missing.length > 5 ? `\\n...и ещё ${{missing.length - 5}} строк(и)` : '';
        if (!confirm(`В таблице есть ${{missing.length}} строк(и) без количества:\\n- ${{preview.join('\\n- ')}}${{tail}}\\n\\nСкачать всё равно?`)) return;
      }}
      const workbook = XLSX.utils.book_new();
      const worksheet = XLSX.utils.json_to_sheet(rows.map((item) => ({{ 'Баркод': item.barcode, 'Количество': item.quantity }})));
      XLSX.utils.book_append_sheet(workbook, worksheet, 'Sheet1');
      XLSX.writeFile(workbook, 'template.xlsx');
    }}
    function togglePriorityRows() {{
      const table = document.getElementById('priority-table');
      const button = document.getElementById('priority-toggle-button');
      if (!table || !button) return;
      const expanded = table.dataset.expanded === 'true';
      table.dataset.expanded = expanded ? 'false' : 'true';
      table.classList.toggle('collapsed', expanded);
      button.textContent = expanded ? 'Показать весь список' : 'Свернуть список';
    }}
    function parseSortValue(text) {{
      const normalized = String(text || '').trim().replace(/\\s+/g, ' ');
      const number = Number(normalized.replace(',', '.').replace(/%/g, ''));
      return !Number.isNaN(number) && normalized !== '' ? number : normalized.toLowerCase();
    }}
    function sortTableByColumn(table, columnIndex, direction) {{
      const tbody = table.tBodies[0];
      const rows = Array.from(tbody.rows);
      rows.sort((left, right) => {{
        const l = parseSortValue(left.cells[columnIndex]?.innerText || '');
        const r = parseSortValue(right.cells[columnIndex]?.innerText || '');
        if (l < r) return direction === 'asc' ? -1 : 1;
        if (l > r) return direction === 'asc' ? 1 : -1;
        return 0;
      }});
      rows.forEach((row) => tbody.appendChild(row));
    }}
    function enableTableSorting() {{
      document.querySelectorAll('table').forEach((table) => {{
        const headers = table.querySelectorAll('thead th');
        headers.forEach((header, index) => {{
          header.addEventListener('click', () => {{
            const next = header.dataset.sortDirection === 'asc' ? 'desc' : 'asc';
            headers.forEach((item) => {{ item.dataset.sortDirection = ''; item.classList.remove('sort-asc', 'sort-desc'); }});
            header.dataset.sortDirection = next;
            header.classList.add(next === 'asc' ? 'sort-asc' : 'sort-desc');
            sortTableByColumn(table, index, next);
          }});
        }});
      }});
    }}
    enableTableSorting();
    refreshProductOptions();
  </script>
</body>
</html>
"""


def _render_priority_table(rows: list[SupplyRow]) -> str:
    if not rows:
        return "<p class='muted'>Сейчас нет позиций, которые стоит везти.</p>"
    body = []
    for index, row in enumerate(rows):
        cls = " class='priority-extra-row'" if index >= 30 else ""
        body.append(
            f"<tr{cls}>"
            f"<td>{html.escape(row.sku)}</td>"
            f"<td class='num'>{row.sales_30d}</td>"
            f"<td class='num'>{row.sale_days_30d}</td>"
            f"<td class='num'>{row.seller_stock}</td>"
            f"<td class='num'>{row.wb_stock}</td>"
            f"<td class='num'>{row.ship_qty}</td>"
            f"<td class='num'>{_money(row.price)}</td>"
            f"<td>{html.escape(row.problem)}</td>"
            f"<td>{html.escape(row.action)}</td>"
            "</tr>"
        )
    return (
        f"<table id='priority-table' class='priority-table collapsed' data-expanded='false'><thead><tr>"
        "<th>SKU</th><th class='num'>Продажи 30д</th><th class='num'>Дни продаж</th>"
        "<th class='num'>Остаток</th><th class='num'>WB</th><th class='num'>К отгрузке</th>"
        "<th class='num'>Цена WB</th><th>Что не так</th><th>Что делать</th>"
        "</tr></thead><tbody>"
        + "".join(body)
        + "</tbody></table>"
    )


def _render_watch_table(rows: list[SupplyRow]) -> str:
    if not rows:
        return "<p class='muted'>Все позиции из выборки сейчас подходят под отгрузку.</p>"
    body = []
    for row in rows:
        body.append(
            "<tr>"
            f"<td>{html.escape(row.sku)}</td>"
            f"<td class='num'>{row.sales_30d}</td>"
            f"<td class='num'>{row.seller_stock}</td>"
            f"<td class='num'>{row.wb_stock}</td>"
            f"<td>{html.escape(row.problem)}</td>"
            f"<td>{html.escape(row.action)}</td>"
            "</tr>"
        )
    return (
        "<table><thead><tr><th>SKU</th><th class='num'>Продажи 30д</th><th class='num'>Остаток</th>"
        "<th class='num'>WB</th><th>Что не так</th><th>Что делать</th></tr></thead><tbody>"
        + "".join(body)
        + "</tbody></table>"
    )


def _render_warehouse_row(row: SupplyRow) -> str:
    return (
        f"<tr data-nm-id='{row.nm_id}'>"
        f"<td>{html.escape(row.sku)}</td>"
        f"<td>{row.nm_id}</td>"
        f"<td>{html.escape(row.vendor_code)}</td>"
        f"<td>{html.escape(row.barcode)}</td>"
        f"<td class='print-hide'>{row.seller_stock}</td>"
        f"<td class='print-hide'>{row.wb_stock}</td>"
        f"<td class='print-hide'>{_money(row.price)}</td>"
        f"<td class='print-hide'>{row.ship_qty}</td>"
        f"<td class='editable' contenteditable='true'>{row.ship_qty}</td>"
        "<td class='editable' contenteditable='true'></td>"
        "<td class='print-hide'><button class='remove-row-button' type='button' onclick='removeWarehouseItem(this)'>Убрать</button></td>"
        "</tr>"
    )


def _render_catalog_item(row: SupplyRow) -> str:
    return (
        "{"
        f"nmId:'{row.nm_id}',"
        f"sku:'{_js(row.sku)}',"
        f"vendorCode:'{_js(row.vendor_code)}',"
        f"barcode:'{_js(row.barcode)}',"
        f"stock:'{row.seller_stock}',"
        f"wbStock:'{row.wb_stock}',"
        f"price:'{_money(row.price)}',"
        f"shipQty:'{row.ship_qty}'"
        "}"
    )


def _render_markdown(rows: list[SupplyRow]) -> str:
    ship_rows = [row for row in rows if row.ship_qty > 0]
    lines = [
        "# Поставка WB",
        "",
        f"- Дата отчета: `{date.today().isoformat()}`",
        "- Период расчета: последние 30 дней",
        f"- SKU с продажами и остатком: `{len(rows)}`",
        f"- SKU к отгрузке: `{len(ship_rows)}`",
        f"- Всего к отгрузке: `{sum(row.ship_qty for row in ship_rows)}` шт.",
        "",
        "## Приоритет к отгрузке",
        "",
        "| SKU | Артикул WB | Артикул продавца | Баркод | Остаток | WB | Продажи 30д | К отгрузке | Цена WB |",
        "| --- | ---: | --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in ship_rows:
        lines.append(
            f"| {_md(row.sku)} | {row.nm_id} | {_md(row.vendor_code)} | {_md(row.barcode)} | "
            f"{row.seller_stock} | {row.wb_stock} | {row.sales_30d} | {row.ship_qty} | {_money(row.price)} |"
        )
    return "\n".join(lines)


def _write_plan_xlsx(path: Path, rows: list[SupplyRow]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Supply plan"
    headers = [
        "SKU",
        "Артикул WB",
        "Артикул продавца",
        "Баркод",
        "Цена WB после скидки продавца",
        "Продажи 30д",
        "Дни продаж 30д",
        "Средние продажи в день",
        "Остаток продавца",
        "WB-остаток",
        "Покрытие WB, дни",
        "Потребность на 30 дней",
        "Безопасно доступно к отгрузке",
        "К отгрузке",
        "Приоритет",
        "Что не так",
        "Что делать",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.sku,
                row.nm_id,
                row.vendor_code,
                row.barcode,
                row.price,
                row.sales_30d,
                row.sale_days_30d,
                row.avg_daily,
                row.seller_stock,
                row.wb_stock,
                row.coverage_days,
                row.demand_30d,
                row.safe_available,
                row.ship_qty,
                row.priority,
                row.problem,
                row.action,
            ]
        )
    _autosize(sheet)
    workbook.save(path)


def _write_template_xlsx(path: Path, rows: list[SupplyRow]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Баркод", "Количество"])
    for row in rows:
        sheet.append([row.barcode, row.ship_qty])
    _autosize(sheet)
    workbook.save(path)


def _autosize(sheet: Any) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(max(width, 12), 70)


def _money(value: float | None) -> str:
    return "" if value is None else f"{value:.2f}"


def _js(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'").replace("\n", " ")


def _md(value: str) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


if __name__ == "__main__":
    main()
